from datetime import datetime
from datetime import date, time
from app.utils.time import now_china
import csv
import io
import re
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, Callable
from collections import defaultdict

from openpyxl import load_workbook, Workbook
from pydantic import ValidationError

from fastapi import APIRouter, Depends, HTTPException, Path, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, exists

from app.api.auth import User, get_current_user, require_roles
from app.api.deps import DbSession
from app.models import (
    AuditLog,
    Inventory,
    InventoryLog,
    Order,
    OrderResource,
    OrderStatusHistory,
    Price,
    Sku,
    Product,
    Channel,
    Spu,
    SkuChannel,
    Supplier,
    Resource,
    ResourceComposition,
    SupplierResource,
    ResourceInventory,
    ResourceInventoryLog,
)
from app.schemas.common import ListResponse, Pagination
from app.schemas.order import OrderCreate, OrderDecision, OrderRead, OrderResourceCreate

router = APIRouter()
COMBINATION_RESOURCE_TYPE = "组合"

STATUS_FIELDS = [
    "is_paid",
    "paid_qty",
    "paid_amount",
    "paid_at",
    "is_issued",
    "issued_qty",
    "issued_amount",
    "issued_at",
    "is_verified",
    "verified_qty",
    "verified_amount",
    "verified_at",
    "is_reserved",
    "reserved_qty",
    "reserved_amount",
    "reserved_at",
    "is_refund_unverified",
    "refund_unverified_qty",
    "refund_unverified_amount",
    "refund_unverified_at",
    "is_refund_unreserved",
    "refund_unreserved_qty",
    "refund_unreserved_amount",
    "refund_unreserved_at",
    "is_refund_verified",
    "refund_verified_qty",
    "refund_verified_amount",
    "refund_verified_at",
    "is_refund_reserved",
    "refund_reserved_qty",
    "refund_reserved_amount",
    "refund_reserved_at",
    "is_completed",
    "completed_qty",
    "completed_amount",
    "completed_at",
    "is_disputed",
    "disputed_qty",
    "disputed_amount",
    "disputed_at",
]

RESOURCE_STATUS_FIELDS = [
    "is_issued",
    "issued_qty",
    "issued_amount",
    "issued_at",
    "issued_remark",
    "is_verified",
    "verified_qty",
    "verified_amount",
    "verified_at",
    "verified_remark",
    "is_refund_unverified",
    "refund_unverified_qty",
    "refund_unverified_amount",
    "refund_unverified_at",
    "refund_unverified_remark",
    "is_refund_unreserved",
    "refund_unreserved_qty",
    "refund_unreserved_amount",
    "refund_unreserved_at",
    "refund_unreserved_remark",
    "is_refund_verified",
    "refund_verified_qty",
    "refund_verified_amount",
    "refund_verified_at",
    "refund_verified_remark",
    "is_refund_reserved",
    "refund_reserved_qty",
    "refund_reserved_amount",
    "refund_reserved_at",
    "refund_reserved_remark",
    "is_completed",
    "completed_qty",
    "completed_amount",
    "completed_at",
    "completed_remark",
    "is_disputed",
    "disputed_qty",
    "disputed_amount",
    "disputed_at",
    "disputed_remark",
    "is_mid_disputed",
    "mid_disputed_qty",
    "mid_disputed_amount",
    "mid_disputed_at",
    "mid_disputed_remark",
]

RESOURCE_STATUS_FLAG_FIELDS = [field for field in RESOURCE_STATUS_FIELDS if field.startswith("is_")]

STATUS_IMPORT_META = {
    "issued": {"label": "是否出票/预约（资源可用）", "time_label": "出票/预约时间"},
    "verified": {"label": "是否核销/消耗", "time_label": "核销/消耗时间"},
    "refund_unverified": {"label": "是否支付后未核销/消耗全部退款", "time_label": "未核销/消耗退款时间"},
    "refund_unreserved": {"label": "是否支付后未出票/预约全部退款", "time_label": "未出票/预约退款时间"},
    "refund_verified": {"label": "是否支付后已核销/消耗全部/部分退款", "time_label": "已核销/消耗退款时间"},
    "refund_reserved": {"label": "是否支付后已出票/预约全部/部分退款", "time_label": "已出票/预约退款时间"},
    "completed": {"label": "是否完成", "time_label": "完成时间"},
    "disputed": {"label": "是否完成后订单产生纠纷", "time_label": "纠纷时间"},
}

STATUS_IMPORT_KEYS = list(STATUS_IMPORT_META.keys())

STATUS_TEMPLATE_FIELDS = [
    ("issued", "出票/预约数量", "出票/预约金额"),
    ("verified", "核销/消耗数量", "核销/消耗金额"),
    ("refund_unverified", "未核销/消耗退款数量", "未核销/消耗退款金额"),
    ("refund_unreserved", "未出票/预约退款数量", "未出票/预约退款金额"),
    ("refund_verified", "已核销/消耗退款数量", "已核销/消耗退款金额"),
    ("refund_reserved", "已出票/预约退款数量", "已出票/预约退款金额"),
    ("completed", "完成数量", "完成金额"),
    ("disputed", "纠纷数量", "纠纷金额"),
]

STATUS_META = {
    "paid": {"label": "支付", "time_label": "支付时间", "amount_label": "支付金额"},
    "issued": {"label": "出票/预约（资源可用）", "time_label": "出票/预约时间", "amount_label": "出票/预约金额"},
    "issue": {"label": "出票/预约", "time_label": "出票/预约时间", "amount_label": "出票/预约金额"},
    "unissue": {"label": "取消出票/预约", "time_label": "取消出票/预约时间", "amount_label": "出票/预约金额"},
    "verified": {"label": "核销/消耗", "time_label": "核销/消耗时间", "amount_label": "核销/消耗金额"},
    "unverify": {"label": "取消核销/消耗", "time_label": "取消核销/消耗时间", "amount_label": "核销/消耗金额"},
    "reserved": {"label": "出票/预约（资源可用）", "time_label": "出票/预约时间", "amount_label": "出票/预约金额"},
    "refund_unverified": {"label": "未核销/消耗退款", "time_label": "未核销/消耗退款时间", "amount_label": "未核销/消耗退款金额"},
    "refund_unreserved": {"label": "未出票/预约退款", "time_label": "未出票/预约退款时间", "amount_label": "未出票/预约退款金额"},
    "refund_verified": {"label": "已核销/消耗退款", "time_label": "已核销/消耗退款时间", "amount_label": "已核销/消耗退款金额"},
    "refund_reserved": {"label": "已出票/预约退款", "time_label": "已出票/预约退款时间", "amount_label": "已出票/预约退款金额"},
    "completed": {"label": "完成", "time_label": "完成时间", "amount_label": "完成金额"},
    "disputed": {"label": "纠纷", "time_label": "纠纷时间", "amount_label": "纠纷金额"},
    "mid_disputed": {"label": "中途纠纷", "time_label": "中途纠纷时间", "amount_label": "中途纠纷金额"},
}

STATUS_RULES = [
    ("paid", "is_paid", "paid_qty", "paid_amount", "paid_at"),
    ("issued", "is_issued", "issued_qty", "issued_amount", "issued_at"),
    ("verified", "is_verified", "verified_qty", "verified_amount", "verified_at"),
    ("reserved", "is_reserved", "reserved_qty", "reserved_amount", "reserved_at"),
    ("refund_unverified", "is_refund_unverified", "refund_unverified_qty", "refund_unverified_amount", "refund_unverified_at"),
    ("refund_unreserved", "is_refund_unreserved", "refund_unreserved_qty", "refund_unreserved_amount", "refund_unreserved_at"),
    ("refund_verified", "is_refund_verified", "refund_verified_qty", "refund_verified_amount", "refund_verified_at"),
    ("refund_reserved", "is_refund_reserved", "refund_reserved_qty", "refund_reserved_amount", "refund_reserved_at"),
    ("completed", "is_completed", "completed_qty", "completed_amount", "completed_at"),
    ("disputed", "is_disputed", "disputed_qty", "disputed_amount", "disputed_at"),
]

RESOURCE_STATUS_RULES = [
    ("issued", "is_issued", "issued_qty", "issued_amount", "issued_at"),
    ("verified", "is_verified", "verified_qty", "verified_amount", "verified_at"),
    ("refund_unverified", "is_refund_unverified", "refund_unverified_qty", "refund_unverified_amount", "refund_unverified_at"),
    ("refund_unreserved", "is_refund_unreserved", "refund_unreserved_qty", "refund_unreserved_amount", "refund_unreserved_at"),
    ("refund_verified", "is_refund_verified", "refund_verified_qty", "refund_verified_amount", "refund_verified_at"),
    ("refund_reserved", "is_refund_reserved", "refund_reserved_qty", "refund_reserved_amount", "refund_reserved_at"),
    ("completed", "is_completed", "completed_qty", "completed_amount", "completed_at"),
    ("disputed", "is_disputed", "disputed_qty", "disputed_amount", "disputed_at"),
    ("mid_disputed", "is_mid_disputed", "mid_disputed_qty", "mid_disputed_amount", "mid_disputed_at"),
]


def _to_decimal(value: Optional[object]) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _require_time_if_yes(key: str, is_flag: bool, at_value: Optional[datetime]):
    if not is_flag:
        return
    meta = STATUS_META.get(key, {})
    time_label = meta.get("time_label", "时间")
    if not at_value:
        raise HTTPException(status_code=400, detail=f"{meta.get('label', key)}选择“是”时必须填写{time_label}")


def _validate_qty_range(key: str, qty: Optional[int], total_qty: int):
    if qty is None:
        return
    meta = STATUS_META.get(key, {})
    if qty < 0:
        raise HTTPException(status_code=400, detail=f"{meta.get('label', key)}数量不能小于 0")
    if qty > total_qty:
        raise HTTPException(status_code=400, detail=f"{meta.get('label', key)}数量不能超过订单数量")


def _validate_amount_range(key: str, amount: Optional[Decimal], max_amount: Optional[Decimal], max_label: str):
    if amount is None:
        return
    meta = STATUS_META.get(key, {})
    if amount < 0:
        raise HTTPException(status_code=400, detail=f"{meta.get('label', key)}金额不能小于 0")
    if max_amount is not None and amount > max_amount:
        raise HTTPException(status_code=400, detail=f"{meta.get('label', key)}金额不能超过{max_label}")


def _effective_qty(is_flag: bool, qty: Optional[int], total_qty: int) -> int:
    if qty is not None:
        return qty
    return total_qty if is_flag else 0


def _validate_status_fields_for_create(payload: OrderCreate, sale_amount: Decimal):
    total_qty = payload.quantity
    paid_amount = _to_decimal(payload.paid_amount)
    issued_amount = _to_decimal(payload.issued_amount)
    verified_amount = _to_decimal(payload.verified_amount)
    reserved_amount = _to_decimal(payload.reserved_amount)

    amount_limits: dict[str, tuple[Optional[Decimal], str]] = {
        "paid": (sale_amount, "销售金额"),
        "issued": (paid_amount or sale_amount, "支付金额" if paid_amount else "销售金额"),
        "verified": (paid_amount or sale_amount, "支付金额" if paid_amount else "销售金额"),
        "reserved": (paid_amount or sale_amount, "支付金额" if paid_amount else "销售金额"),
        "completed": (sale_amount, "销售金额"),
        "disputed": (sale_amount, "销售金额"),
        "refund_unverified": (paid_amount or sale_amount, "支付金额" if paid_amount else "销售金额"),
        "refund_unreserved": (paid_amount or sale_amount, "支付金额" if paid_amount else "销售金额"),
        "refund_verified": (verified_amount or sale_amount, "核销金额" if verified_amount else "销售金额"),
        "refund_reserved": (issued_amount or reserved_amount or sale_amount, "出票/预约金额" if (issued_amount or reserved_amount) else "销售金额"),
    }

    for key, is_field, qty_field, amount_field, at_field in STATUS_RULES:
        is_flag = bool(getattr(payload, is_field))
        _validate_qty_range(key, getattr(payload, qty_field, None), total_qty)
        amount_val = _to_decimal(getattr(payload, amount_field, None))
        max_amount, max_label = amount_limits.get(key, (sale_amount, "销售金额"))
        _validate_amount_range(key, amount_val, max_amount, max_label)

    issued_qty = _effective_qty(payload.is_issued, payload.issued_qty, total_qty)
    verified_qty = _effective_qty(payload.is_verified, payload.verified_qty, total_qty)
    reserved_qty = _effective_qty(payload.is_reserved, payload.reserved_qty, total_qty)
    effective_reserved_qty = reserved_qty if reserved_qty > 0 else issued_qty
    refund_verified_qty = _effective_qty(payload.is_refund_verified, payload.refund_verified_qty, total_qty)
    refund_reserved_qty = _effective_qty(payload.is_refund_reserved, payload.refund_reserved_qty, total_qty)
    refund_unverified_qty = _effective_qty(payload.is_refund_unverified, payload.refund_unverified_qty, total_qty)
    refund_unreserved_qty = _effective_qty(payload.is_refund_unreserved, payload.refund_unreserved_qty, total_qty)

    if refund_verified_qty > 0:
        _ensure_refund_limit(refund_verified_qty, verified_qty, "核销/消耗")
    if refund_reserved_qty > 0:
        _ensure_refund_limit(refund_reserved_qty, effective_reserved_qty, "出票/预约")
    if refund_unverified_qty > 0:
        _ensure_unprocessed_refund_limit(refund_unverified_qty, max(0, total_qty - verified_qty), "核销/消耗")
    if refund_unreserved_qty > 0:
        _ensure_unprocessed_refund_limit(refund_unreserved_qty, max(0, total_qty - effective_reserved_qty), "出票/预约")


def _validate_status_fields_for_resource(payload: object, sale_amount: Decimal, total_qty: int):
    issued_amount = _to_decimal(getattr(payload, "issued_amount", None))
    verified_amount = _to_decimal(getattr(payload, "verified_amount", None))
    reserved_amount = _to_decimal(getattr(payload, "reserved_amount", None))

    amount_limits: dict[str, tuple[Optional[Decimal], str]] = {
        "issued": (sale_amount, "销售金额"),
        "verified": (sale_amount, "销售金额"),
        "reserved": (sale_amount, "销售金额"),
        "completed": (sale_amount, "销售金额"),
        "disputed": (sale_amount, "销售金额"),
        "mid_disputed": (sale_amount, "销售金额"),
        "refund_unverified": (sale_amount, "销售金额"),
        "refund_unreserved": (sale_amount, "销售金额"),
        "refund_verified": (verified_amount or sale_amount, "核销金额" if verified_amount else "销售金额"),
        "refund_reserved": (issued_amount or reserved_amount or sale_amount, "出票/预约金额" if (issued_amount or reserved_amount) else "销售金额"),
    }

    for key, is_field, qty_field, amount_field, at_field in RESOURCE_STATUS_RULES:
        is_flag = bool(getattr(payload, is_field))
        _validate_qty_range(key, getattr(payload, qty_field, None), total_qty)
        amount_val = _to_decimal(getattr(payload, amount_field, None))
        max_amount, max_label = amount_limits.get(key, (sale_amount, "销售金额"))
        _validate_amount_range(key, amount_val, max_amount, max_label)

    issued_qty = _effective_qty(getattr(payload, "is_issued", False), getattr(payload, "issued_qty", None), total_qty)
    verified_qty = _effective_qty(getattr(payload, "is_verified", False), getattr(payload, "verified_qty", None), total_qty)
    reserved_qty = _effective_qty(getattr(payload, "is_reserved", False), getattr(payload, "reserved_qty", None), total_qty)
    effective_reserved_qty = reserved_qty if reserved_qty > 0 else issued_qty
    refund_verified_qty = _effective_qty(
        getattr(payload, "is_refund_verified", False),
        getattr(payload, "refund_verified_qty", None),
        total_qty,
    )
    refund_reserved_qty = _effective_qty(
        getattr(payload, "is_refund_reserved", False),
        getattr(payload, "refund_reserved_qty", None),
        total_qty,
    )
    refund_unverified_qty = _effective_qty(
        getattr(payload, "is_refund_unverified", False),
        getattr(payload, "refund_unverified_qty", None),
        total_qty,
    )
    refund_unreserved_qty = _effective_qty(
        getattr(payload, "is_refund_unreserved", False),
        getattr(payload, "refund_unreserved_qty", None),
        total_qty,
    )

    if refund_verified_qty > 0:
        _ensure_refund_limit(refund_verified_qty, verified_qty, "核销/消耗")
    if refund_reserved_qty > 0:
        _ensure_refund_limit(refund_reserved_qty, effective_reserved_qty, "出票/预约")
    if refund_unverified_qty > 0:
        _ensure_unprocessed_refund_limit(refund_unverified_qty, max(0, total_qty - verified_qty), "核销/消耗")
    if refund_unreserved_qty > 0:
        _ensure_unprocessed_refund_limit(refund_unreserved_qty, max(0, total_qty - effective_reserved_qty), "出票/预约")


def _normalize_header(value: str) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"[\s\u3000]+", "", text)
    text = re.sub(r"[()（）\[\]【】]", "", text)
    text = re.sub(r"[\\/|·•,，。:：;；、\-–—_]+", "", text)
    return text


def _normalize_text(value: str) -> str:
    return str(value or "").strip().lower()


def _build_header_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}

    def add(key: str, canonical: str):
        aliases[_normalize_header(key)] = canonical

    # Base fields
    add("订单号", "order_no")
    add("订单编号", "order_no")
    add("order_no", "order_no")
    add("orderno", "order_no")

    add("渠道", "channel")
    add("渠道名称", "channel_name")
    add("channel", "channel")
    add("channel_name", "channel_name")
    add("渠道id", "channel_id")
    add("渠道编号", "channel_id")
    add("channel_id", "channel_id")
    add("channelid", "channel_id")

    add("spu", "spu")
    add("spu_id", "spu_id")
    add("spuid", "spu_id")
    add("spu名称", "spu_name")
    add("spu_name", "spu_name")
    add("spu编码", "spu_code")
    add("spu_code", "spu_code")

    add("sku", "sku")
    add("sku_id", "sku_id")
    add("skuid", "sku_id")
    add("sku名称", "sku_name")
    add("sku_name", "sku_name")

    add("数量", "quantity")
    add("qty", "quantity")
    add("quantity", "quantity")

    add("销售金额", "sale_amount")
    add("销售额", "sale_amount")
    add("成交金额", "sale_amount")
    add("售价", "sale_amount")
    add("sale_amount", "sale_amount")
    add("sales_amount", "sale_amount")
    add("销售单价", "sale_price")
    add("单价", "sale_price")
    add("sale_price", "sale_price")
    add("saleprice", "sale_price")

    add("出行日期", "travel_date")
    add("出行时间", "travel_date")
    add("travel_date", "travel_date")
    add("traveldate", "travel_date")

    add("支付时间", "paid_at")
    add("支付日期", "paid_at")
    add("paid_at", "paid_at")
    add("paidtime", "paid_at")
    add("支付数量", "paid_qty")
    add("支付金额", "paid_amount")

    add("备注", "remark")
    add("订单备注", "remark")
    add("remark", "remark")

    # Status aliases
    add("????/??/???", "is_issued")
    add("????(??)/??????/?????", "is_issued")
    add("????/????????", "is_issued")
    add("????/??", "is_issued")
    add("????", "is_issued")
    add("????", "is_issued")
    add("??????", "is_issued")
    add("?????", "is_issued")
    add("????", "is_issued")
    add("????", "issued_qty")
    add("????", "issued_qty")
    add("?????", "issued_qty")
    add("????", "issued_qty")
    add("??/????", "issued_qty")
    add("????", "issued_amount")
    add("????", "issued_amount")
    add("?????", "issued_amount")
    add("????", "issued_amount")
    add("??/????", "issued_amount")
    add("????", "issued_at")
    add("????", "issued_at")
    add("?????", "issued_at")
    add("????", "issued_at")
    add("??/????", "issued_at")

    add("????", "is_verified")
    add("????/??", "is_verified")
    add("????", "verified_qty")
    add("??/????", "verified_qty")
    add("????", "verified_amount")
    add("??/????", "verified_amount")
    add("????", "verified_at")
    add("??/????", "verified_at")

    add("????????????", "is_refund_unverified")
    add("????????/??????", "is_refund_unverified")
    add("?????", "is_refund_unverified")
    add("???/????", "is_refund_unverified")
    add("???????", "refund_unverified_qty")
    add("???/??????", "refund_unverified_qty")
    add("???????", "refund_unverified_amount")
    add("???/??????", "refund_unverified_amount")
    add("???????", "refund_unverified_at")
    add("???/??????", "refund_unverified_at")

    add("????????????", "is_refund_unreserved")
    add("????????/??????", "is_refund_unreserved")
    add("?????", "is_refund_unreserved")
    add("???/????", "is_refund_unreserved")
    add("???????", "refund_unreserved_qty")
    add("???/??????", "refund_unreserved_qty")
    add("???????", "refund_unreserved_amount")
    add("???/??????", "refund_unreserved_amount")
    add("???????", "refund_unreserved_at")
    add("???/??????", "refund_unreserved_at")

    add("??????????/????", "is_refund_verified")
    add("????????/????/????", "is_refund_verified")
    add("?????", "is_refund_verified")
    add("???/????", "is_refund_verified")
    add("???????", "refund_verified_qty")
    add("???/??????", "refund_verified_qty")
    add("???????", "refund_verified_amount")
    add("???/??????", "refund_verified_amount")
    add("???????", "refund_verified_at")
    add("???/??????", "refund_verified_at")

    add("??????????/????", "is_refund_reserved")
    add("????????/????/????", "is_refund_reserved")
    add("?????", "is_refund_reserved")
    add("???/????", "is_refund_reserved")
    add("???????", "refund_reserved_qty")
    add("???/??????", "refund_reserved_qty")
    add("???????", "refund_reserved_amount")
    add("???/??????", "refund_reserved_amount")
    add("???????", "refund_reserved_at")
    add("???/??????", "refund_reserved_at")

    add("????", "is_completed")
    add("????", "completed_qty")
    add("????", "completed_amount")
    add("????", "completed_at")

    add("???????????", "is_disputed")
    add("????", "disputed_qty")
    add("????", "disputed_amount")
    add("????", "disputed_at")

    # Canonical status keys
    for key in STATUS_FIELDS:
        add(key, key)

    return aliases


HEADER_ALIASES = _build_header_aliases()


def _parse_bool(value) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "是", "已", "对"}:
        return True
    if text in {"0", "false", "no", "n", "否", "未", "错"}:
        return False
    return None


def _parse_int(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)
    try:
        num = float(text)
        if num.is_integer():
            return int(num)
    except ValueError:
        return None
    return None


def _parse_decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("¥", "").replace("元", "")
    try:
        return Decimal(text)
    except Exception:
        return None


def _parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                return datetime.combine(parsed.date(), time.min)
            return parsed
        except ValueError:
            continue
    return None


def _normalize_row(raw: dict) -> dict:
    normalized: dict[str, object] = {}
    for key, value in raw.items():
        if key is None:
            continue
        canonical = HEADER_ALIASES.get(_normalize_header(key))
        if not canonical:
            continue
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        if canonical not in normalized or normalized.get(canonical) in (None, ""):
            normalized[canonical] = value
    return normalized


async def _read_upload_rows(file: UploadFile) -> list[tuple[int, dict]]:
    content = await file.read()
    if not content:
        return []
    filename = (file.filename or "").lower()
    rows: list[tuple[int, dict]] = []

    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("gbk", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for idx, row in enumerate(reader, start=2):
            if not row or all(v is None or (isinstance(v, str) and not v.strip()) for v in row.values()):
                continue
            rows.append((idx, row))
        return rows

    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), data_only=True)
        sheet = wb.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            return []
        headers = [str(c).strip() if c is not None else "" for c in raw_rows[0]]
        for idx, values in enumerate(raw_rows[1:], start=2):
            if values is None:
                continue
            row = {}
            empty = True
            for h, v in zip(headers, values):
                if not h:
                    continue
                row[h] = v
                if v is not None and (not isinstance(v, str) or v.strip()):
                    empty = False
            if empty:
                continue
            rows.append((idx, row))
        return rows

    raise HTTPException(status_code=400, detail="不支持的文件类型")

def _calc_amounts(sale_price: Decimal, cost_price: Decimal | None, qty: int):
    sale_amount = sale_price * qty
    cost_amount = cost_price * qty if cost_price is not None else None
    profit_amount = sale_amount - cost_amount if cost_amount is not None else None
    return sale_amount, cost_amount, profit_amount


async def _active_price(db: DbSession, sku_id: int, channel_id: int, travel_date):
    stmt = select(Price).where(
        Price.sku_id == sku_id,
        Price.channel_id == channel_id,
        Price.status == "active",
        Price.start_at <= travel_date,
        Price.end_at >= travel_date,
    )
    return await db.scalar(stmt)


async def _calc_sku_channel_total(db: DbSession, sku_id: int, channel_id: int, target_date) -> int:
    from app.models import ProductResource, ResourceInventory

    sku = await db.get(Sku, sku_id)
    if not sku:
        raise HTTPException(status_code=404, detail="SKU 不存在")
    product = await db.get(Product, sku.product_id)
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")

    ratio = 0
    if product.allowed_channels:
        for alloc in product.allowed_channels:
            if isinstance(alloc, dict):
                if alloc.get("channel_id") == channel_id:
                    ratio = alloc.get("stock_ratio", 0) or 0
                    if alloc.get("stock_ratio") is None:
                        ratio = 100
                    break
            else:
                try:
                    cid = int(alloc)
                except (TypeError, ValueError):
                    continue
                if cid == channel_id:
                    ratio = 100
                    break
    if ratio <= 0:
        return 0

    resources_stmt = select(ProductResource).where(ProductResource.product_id == product.id)
    product_resources = list(await db.scalars(resources_stmt))
    required_resources = [pr for pr in product_resources if pr.required_flag]
    if not required_resources:
        return 0

    resource_ids = [pr.resource_id for pr in required_resources]
    inv_stmt = (
        select(ResourceInventory, SupplierResource.resource_id, SupplierResource.supplier_id)
        .join(SupplierResource)
        .where(
            SupplierResource.resource_id.in_(resource_ids),
            ResourceInventory.inventory_date == target_date,
        )
    )
    inventory_rows = (await db.execute(inv_stmt)).all()

    detailed_lookup: dict[tuple[int, int, str], int] = {}
    total_lookup: dict[tuple[int, str], int] = {}
    date_str = str(target_date)
    for inv, r_id, s_id in inventory_rows:
        available = max(0, inv.total_qty - inv.sold_qty - inv.frozen_qty)
        detailed_lookup[(r_id, s_id, date_str)] = detailed_lookup.get((r_id, s_id, date_str), 0) + available
        total_lookup[(r_id, date_str)] = total_lookup.get((r_id, date_str), 0) + available

    min_qty = None
    for pr in required_resources:
        if pr.supplier_mode == "locked" and pr.supplier_ids:
            resource_available = sum(detailed_lookup.get((pr.resource_id, sid, date_str), 0) for sid in pr.supplier_ids)
        else:
            resource_available = total_lookup.get((pr.resource_id, date_str), 0)
        qty_from_resource = resource_available // pr.quantity if pr.quantity > 0 else 0
        min_qty = qty_from_resource if min_qty is None else min(min_qty, qty_from_resource)

    product_available = min_qty if min_qty is not None else 0
    return int(product_available * (ratio / 100))


async def _freeze_inventory(db: DbSession, sku_id: int, channel_id: int, travel_date, qty: int, operator: str, order_id: Optional[int]):
    inv = await db.scalar(
        select(Inventory).where(Inventory.sku_id == sku_id, Inventory.inventory_date == travel_date).with_for_update()
    )
    computed_available = await _calc_sku_channel_total(db, sku_id, channel_id, travel_date)
    if inv:
        before = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
        inv.total_qty = inv.sold_qty + inv.frozen_qty + computed_available
    else:
        inv = Inventory(
            sku_id=sku_id,
            inventory_date=travel_date,
            total_qty=computed_available,
            frozen_qty=0,
            sold_qty=0,
            status="normal",
        )
        before = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
        db.add(inv)

    available = inv.total_qty - inv.frozen_qty - inv.sold_qty
    if available < qty:
        raise HTTPException(status_code=400, detail="SKU 库存不足，无法下单")

    inv.frozen_qty += qty
    inv.updated_at = now_china()
    after = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    log = InventoryLog(
        sku_id=sku_id,
        inventory_date=travel_date,
        change_type="freeze",
        before_qty=before,
        after_qty=after,
        related_order_id=order_id,
        operator=operator,
        operated_at=now_china(),
    )
    db.add_all([inv, log])


async def _consume_inventory(db: DbSession, sku_id: int, travel_date, qty: int, operator: str, order_id: int, action: str):
    inv = await db.scalar(
        select(Inventory).where(Inventory.sku_id == sku_id, Inventory.inventory_date == travel_date).with_for_update()
    )
    if not inv:
        raise HTTPException(status_code=400, detail="SKU 库存未初始化")
    if inv.frozen_qty < qty:
        raise HTTPException(status_code=400, detail="SKU 冻结库存不足")
    before = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    inv.frozen_qty -= qty
    if action == "verify":
        inv.sold_qty += qty
    inv.updated_at = now_china()
    after = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    log = InventoryLog(
        sku_id=sku_id,
        inventory_date=travel_date,
        change_type="consume" if action == "verify" else "release",
        before_qty=before,
        after_qty=after,
        related_order_id=order_id,
        operator=operator,
        operated_at=now_china(),
    )
    db.add_all([inv, log])


async def _unconsume_inventory(db: DbSession, sku_id: int, travel_date, qty: int, operator: str, order_id: int):
    inv = await db.scalar(
        select(Inventory).where(Inventory.sku_id == sku_id, Inventory.inventory_date == travel_date).with_for_update()
    )
    if not inv:
        raise HTTPException(status_code=400, detail="SKU 库存未初始化")
    if inv.sold_qty < qty:
        raise HTTPException(status_code=400, detail="SKU 已售库存不足")
    before = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    inv.sold_qty -= qty
    inv.frozen_qty += qty
    inv.updated_at = now_china()
    after = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    log = InventoryLog(
        sku_id=sku_id,
        inventory_date=travel_date,
        change_type="unconsume",
        before_qty=before,
        after_qty=after,
        related_order_id=order_id,
        operator=operator,
        operated_at=now_china(),
    )
    db.add_all([inv, log])


async def _return_inventory(db: DbSession, sku_id: int, travel_date, qty: int, operator: str, order_id: int):
    inv = await db.scalar(
        select(Inventory).where(Inventory.sku_id == sku_id, Inventory.inventory_date == travel_date).with_for_update()
    )
    if not inv:
        raise HTTPException(status_code=400, detail="SKU 库存未初始化")
    if inv.sold_qty < qty:
        raise HTTPException(status_code=400, detail="SKU 已售库存不足")
    before = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    inv.sold_qty -= qty
    inv.updated_at = now_china()
    after = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
    log = InventoryLog(
        sku_id=sku_id,
        inventory_date=travel_date,
        change_type="return",
        before_qty=before,
        after_qty=after,
        related_order_id=order_id,
        operator=operator,
        operated_at=now_china(),
    )
    db.add_all([inv, log])


def _resolve_qty(order: Order, payload_qty: Optional[int], current_qty: Optional[int]) -> tuple[int, int]:
    target = payload_qty if payload_qty is not None else order.quantity
    if target <= 0:
        raise HTTPException(status_code=400, detail="数量必须大于 0")
    if target > order.quantity:
        raise HTTPException(status_code=400, detail="数量不能超过订单数量")
    applied = current_qty or 0
    if applied < 0:
        applied = 0
    delta = max(0, target - applied)
    return target, delta


def _ensure_refund_limit(target_qty: int, max_allowed: int, label: str):
    if target_qty > max_allowed:
        raise HTTPException(status_code=400, detail=f"{label}退款数量不能超过已{label}数量")


def _ensure_unprocessed_refund_limit(target_qty: int, max_allowed: int, label: str):
    if target_qty > max_allowed:
        raise HTTPException(status_code=400, detail=f"{label}退款数量不能超过未{label}数量")


async def _apply_resource_inventory(
    db: DbSession,
    order: Order,
    qty: int,
    operator: str,
    action: str,
    resource_predicate: Optional[Callable[[OrderResource], bool]] = None,
):
    if qty <= 0:
        return
    rows = await db.execute(
        select(OrderResource, SupplierResource)
        .join(
            SupplierResource,
            (SupplierResource.resource_id == OrderResource.resource_id)
            & (SupplierResource.supplier_id == OrderResource.supplier_id),
        )
        .where(OrderResource.order_id == order.id)
    )
    resources = rows.all()
    if resource_predicate:
        resources = [(order_resource, supplier_resource) for order_resource, supplier_resource in resources if resource_predicate(order_resource)]
    if not resources:
        return

    by_resource: dict[tuple[int, date], list[tuple[OrderResource, SupplierResource]]] = {}
    for order_resource, supplier_resource in resources:
        travel_date = order_resource.travel_date or order.travel_date
        if not travel_date:
            continue
        by_resource.setdefault((order_resource.resource_id, travel_date), []).append((order_resource, supplier_resource))

    for (resource_id, travel_date), items in by_resource.items():
        if order.quantity <= 0:
            continue
        total_qty = sum(or_rec.quantity for or_rec, _ in items)
        if total_qty <= 0:
            continue
        if total_qty % order.quantity != 0:
            raise HTTPException(status_code=400, detail="资源数量与订单数量不匹配")

        per_order_unit = total_qty // order.quantity
        resource_total_needed = per_order_unit * qty
        if resource_total_needed <= 0:
            continue

        allocations = []
        allocated = 0
        for order_resource, supplier_resource in items:
            raw = (Decimal(order_resource.quantity) * Decimal(resource_total_needed)) / Decimal(total_qty)
            base_qty = int(raw)
            frac = raw - base_qty
            allocations.append(
                {
                    "order_resource": order_resource,
                    "supplier_resource": supplier_resource,
                    "qty": base_qty,
                    "frac": frac,
                }
            )
            allocated += base_qty

        remainder = resource_total_needed - allocated
        if remainder > 0 and allocations:
            allocations.sort(
                key=lambda x: (x["frac"], x["supplier_resource"].supplier_id),
                reverse=True,
            )
            for i in range(remainder):
                allocations[i % len(allocations)]["qty"] += 1

        for alloc in allocations:
            resource_qty = alloc["qty"]
            if resource_qty <= 0:
                continue
            supplier_resource = alloc["supplier_resource"]

            inv = await db.scalar(
                select(ResourceInventory)
                .where(
                    ResourceInventory.supplier_resource_id == supplier_resource.id,
                    ResourceInventory.inventory_date == travel_date,
                )
                .with_for_update()
            )
            if not inv:
                raise HTTPException(status_code=400, detail="资源库存未初始化")

            available = inv.total_qty - inv.frozen_qty - inv.sold_qty
            before = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}

            if action == "freeze":
                if available < resource_qty:
                    raise HTTPException(status_code=400, detail="资源库存不足，无法冻结")
                inv.frozen_qty += resource_qty
            elif action == "consume":
                if inv.frozen_qty < resource_qty:
                    raise HTTPException(status_code=400, detail="资源冻结库存不足")
                inv.frozen_qty -= resource_qty
                inv.sold_qty += resource_qty
            elif action == "release":
                if inv.frozen_qty < resource_qty:
                    raise HTTPException(status_code=400, detail="资源冻结库存不足")
                inv.frozen_qty -= resource_qty
            elif action == "return":
                if inv.sold_qty < resource_qty:
                    raise HTTPException(status_code=400, detail="资源已售库存不足")
                inv.sold_qty -= resource_qty
            elif action == "unconsume":
                if inv.sold_qty < resource_qty:
                    raise HTTPException(status_code=400, detail="资源已售库存不足")
                inv.sold_qty -= resource_qty
                inv.frozen_qty += resource_qty
            else:
                raise HTTPException(status_code=400, detail="不支持的资源库存操作")

            inv.updated_at = now_china().isoformat()
            after = {"total": inv.total_qty, "frozen": inv.frozen_qty, "sold": inv.sold_qty}
            log = ResourceInventoryLog(
                supplier_resource_id=supplier_resource.id,
                inventory_date=travel_date,
                change_type=action,
                before_qty=before,
                after_qty=after,
                related_order_id=order.id,
                operator=operator,
                operated_at=now_china(),
            )
            db.add_all([inv, log])


async def _build_combination_snapshots(
    db: DbSession,
    root_resource_ids: list[int],
) -> dict[int, list[dict]]:
    if not root_resource_ids:
        return {}

    root_resources = list(await db.scalars(select(Resource).where(Resource.id.in_(root_resource_ids))))
    root_combo_ids = [r.id for r in root_resources if r.is_combination]
    if not root_combo_ids:
        return {}

    relation_rows = (
        await db.execute(
            select(
                ResourceComposition.parent_resource_id,
                ResourceComposition.child_resource_id,
                ResourceComposition.position,
            ).order_by(
                ResourceComposition.parent_resource_id.asc(),
                ResourceComposition.position.asc(),
                ResourceComposition.id.asc(),
            )
        )
    ).all()

    resource_ids: set[int] = set(root_combo_ids)
    children_map: dict[int, list[int]] = defaultdict(list)
    for parent_id, child_id, _ in relation_rows:
        resource_ids.add(parent_id)
        resource_ids.add(child_id)
        children_map[parent_id].append(child_id)

    resources = list(await db.scalars(select(Resource).where(Resource.id.in_(resource_ids))))
    resource_map = {resource.id: resource for resource in resources}

    def build_member_node(resource_id: int, stack: set[int]) -> Optional[dict]:
        resource = resource_map.get(resource_id)
        if not resource:
            return None
        node = {
            "resource_id": resource.id,
            "resource_name": resource.resource_name,
            "resource_type": COMBINATION_RESOURCE_TYPE if resource.is_combination else resource.resource_type,
            "poi_id": resource.poi_id,
            "is_combination": bool(resource.is_combination),
        }
        if resource.is_combination:
            children: list[dict] = []
            for child_id in children_map.get(resource_id, []):
                if child_id in stack:
                    continue
                child_node = build_member_node(child_id, stack | {resource_id})
                if child_node:
                    children.append(child_node)
            node["members"] = children
        return node

    snapshots: dict[int, list[dict]] = {}
    for root_id in root_combo_ids:
        members: list[dict] = []
        for child_id in children_map.get(root_id, []):
            child_node = build_member_node(child_id, {root_id})
            if child_node:
                members.append(child_node)
        snapshots[root_id] = members
    return snapshots


@router.get("/orders", response_model=ListResponse)
async def list_orders(
    db: DbSession,
    _: User = Depends(get_current_user),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=1000),
    keyword: Optional[str] = Query(default=None),
    spu_id: Optional[int] = Query(default=None),
    sku_id: Optional[int] = Query(default=None),
    channel_id: Optional[int] = Query(default=None),
    travel_date: Optional[date] = Query(default=None),
    paid_at: Optional[datetime] = Query(default=None),
    paid_date: Optional[date] = Query(default=None),
):
    stmt = (
        select(Order, Channel.channel_name, Sku.sku_name, Product.product_name, Sku.spu_id, Spu.name.label("spu_name"))
        .join(Channel, Channel.id == Order.channel_id)
        .join(Sku, Sku.id == Order.sku_id)
        .join(Product, Product.id == Order.product_id)
        .join(Spu, Spu.id == Sku.spu_id)
    )
    if keyword:
        kw = keyword.strip()
        if kw:
            stmt = stmt.where(Order.order_no.ilike(f"%{kw}%"))
    if spu_id:
        stmt = stmt.where(Sku.spu_id == spu_id)
    if sku_id:
        stmt = stmt.where(Order.sku_id == sku_id)
    if channel_id:
        stmt = stmt.where(Order.channel_id == channel_id)
    if travel_date:
        stmt = stmt.where(
            exists(
                select(1).where(
                    OrderResource.order_id == Order.id,
                    OrderResource.travel_date == travel_date,
                )
            )
        )
    if paid_date:
        start_dt = datetime.combine(paid_date, time.min)
        end_dt = datetime.combine(paid_date, time.max)
        stmt = stmt.where(Order.paid_at >= start_dt, Order.paid_at <= end_dt)
    elif paid_at:
        stmt = stmt.where(Order.paid_at == paid_at)
    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    result = await db.execute(
        stmt.order_by(Spu.id.asc(), Order.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )
    rows = result.all()
    order_ids = [order.id for order, *_ in rows]
    resource_summary: dict[int, dict] = {}
    resource_details_map: dict[int, list[dict]] = {}
    if order_ids:
        res_rows = (
            await db.execute(
                select(OrderResource, Resource, Supplier)
                .join(Resource, Resource.id == OrderResource.resource_id)
                .join(Supplier, Supplier.id == OrderResource.supplier_id)
                .where(OrderResource.order_id.in_(order_ids))
                .order_by(OrderResource.order_id.asc(), OrderResource.id.asc())
            )
        ).all()
        for order_resource, resource, supplier in res_rows:
            summary = resource_summary.setdefault(order_resource.order_id, {"travel_date": None, "flags": {}})
            if order_resource.travel_date:
                if summary["travel_date"] is None or order_resource.travel_date < summary["travel_date"]:
                    summary["travel_date"] = order_resource.travel_date
            for flag in RESOURCE_STATUS_FLAG_FIELDS:
                if getattr(order_resource, flag, False):
                    summary["flags"][flag] = True

            detail = {
                "order_resource_id": order_resource.id,
                "resource_id": resource.id,
                "resource_name": resource.resource_name,
                "resource_type": COMBINATION_RESOURCE_TYPE if resource.is_combination else resource.resource_type,
                "is_combination": bool(resource.is_combination),
                "supplier_id": supplier.id,
                "supplier_name": supplier.supplier_name,
                "travel_date": order_resource.travel_date,
                "quantity": order_resource.quantity,
                "settlement_price": float(order_resource.settlement_price) if order_resource.settlement_price is not None else None,
                "cost_amount": float(order_resource.cost_amount) if order_resource.cost_amount is not None else None,
                "composition_snapshot": order_resource.composition_snapshot,
                "composition_snapshot_at": order_resource.composition_snapshot_at,
            }
            for field in RESOURCE_STATUS_FIELDS:
                detail[field] = getattr(order_resource, field, None)
            resource_details_map.setdefault(order_resource.order_id, []).append(detail)

    items = []
    for order, channel_name, sku_name, product_name, spu_id, spu_name in rows:
        payload = OrderRead.model_validate(order).model_dump()
        payload["channel_name"] = channel_name
        payload["sku_name"] = sku_name
        payload["product_name"] = product_name
        payload["spu_id"] = spu_id
        payload["spu_name"] = spu_name
        payload["resource_details"] = resource_details_map.get(order.id, [])
        summary = resource_summary.get(order.id)
        if summary:
            if summary.get("travel_date"):
                payload["travel_date"] = summary["travel_date"]
            flags = summary.get("flags", {})
            for flag in RESOURCE_STATUS_FLAG_FIELDS:
                payload[flag] = bool(flags.get(flag, False))
        items.append(payload)
    return ListResponse(
        items=items,
        pagination=Pagination(total=total or 0, page=page, page_size=page_size),
    )


@router.post("/orders", response_model=OrderRead, status_code=status.HTTP_201_CREATED)
async def create_order(payload: OrderCreate, db: DbSession, user: User = Depends(require_roles(["admin", "operator", "csr"]))):
    # 1. Check duplicate
    dup = await db.scalar(
        select(Order).where(Order.order_no == payload.order_no, Order.channel_id == payload.channel_id)
    )
    if dup:
        raise HTTPException(status_code=400, detail="该渠道下订单号已存在")

    channel = await db.get(Channel, payload.channel_id)
    if not channel:
        raise HTTPException(status_code=404, detail="渠道不存在")

    sku = await db.get(Sku, payload.sku_id)
    if not sku:
        raise HTTPException(status_code=404, detail="SKU 不存在")
    if sku.status != "active":
        raise HTTPException(status_code=400, detail="SKU 未上架，无法下单")

    now_date = now_china().date()
    if sku.sale_start and now_date < sku.sale_start:
        raise HTTPException(status_code=400, detail="SKU 未到售卖期，无法下单")
    if sku.sale_end and now_date > sku.sale_end:
        raise HTTPException(status_code=400, detail="SKU 已过售卖期，无法下单")

    sku_channel = await db.scalar(
        select(SkuChannel).where(
            SkuChannel.sku_id == payload.sku_id,
            SkuChannel.channel_id == payload.channel_id,
            SkuChannel.status == "active",
        )
    )
    if not sku_channel:
        raise HTTPException(status_code=400, detail="SKU 未绑定该渠道或已停用")

    product_id = payload.product_id or sku.product_id
    if payload.product_id and payload.product_id != sku.product_id:
        raise HTTPException(status_code=400, detail="SKU 不属于该产品")

    product = await db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="产品不存在")

    if not product.allowed_channels:
        raise HTTPException(status_code=400, detail="产品未配置渠道库存占比，无法下单")
    channel_ratio = None
    for alloc in product.allowed_channels:
        if isinstance(alloc, dict):
            if alloc.get("channel_id") == payload.channel_id:
                channel_ratio = alloc.get("stock_ratio", 0)
                if alloc.get("stock_ratio") is None:
                    channel_ratio = 100
                break
        else:
            try:
                cid = int(alloc)
            except (TypeError, ValueError):
                continue
            if cid == payload.channel_id:
                channel_ratio = 100
                break
    if channel_ratio is None:
        raise HTTPException(status_code=400, detail="该产品未配置此渠道库存占比")
    if channel_ratio <= 0:
        raise HTTPException(status_code=400, detail="该渠道库存占比为 0，无法下单")
    
    # 2. Process Product Resources & Choose Suppliers
    from app.models import ProductResource
    
    # Get product resources
    pres = await db.scalars(select(ProductResource).where(ProductResource.product_id == product_id))
    pres = list(pres)
    
    total_cost = Decimal("0.00")
    order_resources_data = []  # List of dicts to create OrderResource

    # Manual selection map: resource_id -> supplier_id
    selections = payload.resource_selections or {}

    resource_input_map: dict[int, OrderResourceCreate] = {}
    if payload.resource_items:
        for item in payload.resource_items:
            if item.resource_id in resource_input_map:
                raise HTTPException(status_code=400, detail=f"资源 {item.resource_id} 重复设置")
            resource_input_map[item.resource_id] = item
        valid_resource_ids = {line.resource_id for line in pres}
        unknown_ids = [rid for rid in resource_input_map.keys() if rid not in valid_resource_ids]
        if unknown_ids:
            raise HTTPException(status_code=400, detail=f"资源 {unknown_ids[0]} 不属于该产品")
    else:
        if not payload.travel_date:
            raise HTTPException(status_code=400, detail="出行日期不能为空")
        default_status = {}
        for field in RESOURCE_STATUS_FIELDS:
            value = getattr(payload, field, None)
            if field.startswith("is_") and value is None:
                value = False
            default_status[field] = value
        for line in pres:
            resource_input_map[line.resource_id] = OrderResourceCreate(
                resource_id=line.resource_id,
                travel_date=payload.travel_date,
                **default_status,
            )

    if payload.paid_at is None:
        raise HTTPException(status_code=400, detail="支付时间不能为空")
    for item in resource_input_map.values():
        if getattr(item, "is_verified", False) and not getattr(item, "is_issued", False):
            raise HTTPException(status_code=400, detail="核销/消耗前必须先出票/预约")

    resource_dates = [item.travel_date for item in resource_input_map.values() if item.travel_date]
    if not resource_dates and payload.travel_date:
        resource_dates = [payload.travel_date]
    if sku.travel_start and any(d < sku.travel_start for d in resource_dates):
        raise HTTPException(status_code=400, detail="出行日期早于可用范围")
    if sku.travel_end and any(d > sku.travel_end for d in resource_dates):
        raise HTTPException(status_code=400, detail="出行日期晚于可用范围")
    primary_travel_date = min(resource_dates) if resource_dates else None
    combination_snapshots = await _build_combination_snapshots(
        db,
        [line.resource_id for line in pres],
    )

    for line in pres:
        qty_needed = line.quantity * payload.quantity
        if qty_needed <= 0:
            continue

        resource_input = resource_input_map.get(line.resource_id)
        if not resource_input:
            raise HTTPException(status_code=400, detail=f"资源 {line.resource_id} 未设置出行日期")
        travel_date = resource_input.travel_date
        if not travel_date:
            raise HTTPException(status_code=400, detail="资源出行日期不能为空")
        composition_snapshot = combination_snapshots.get(line.resource_id)
            
        # Determine candidate suppliers
        candidates = []
        if line.supplier_mode == 'locked':
            if not line.supplier_ids:
                raise HTTPException(status_code=400, detail=f"资源 {line.resource_id} 已锁定供应商但未配置供应商")
            # Fetch SupplierResources for these IDs
            stmt = select(SupplierResource).where(
                SupplierResource.resource_id == line.resource_id,
                SupplierResource.supplier_id.in_(line.supplier_ids),
                SupplierResource.supply_status == "active",
            )
            # candidates = list(await db.scalars(stmt))
            candidates = (await db.scalars(stmt)).all()
        else: # 'auto'
            # Fetch ALL SupplierResources for this resource
            stmt = select(SupplierResource).where(
                SupplierResource.resource_id == line.resource_id,
                SupplierResource.supply_status == "active",
            )
            # candidates = list(await db.scalars(stmt))
            candidates = (await db.scalars(stmt)).all()
            
        if not candidates:
            raise HTTPException(status_code=400, detail=f"资源 {line.resource_id} 没有可用供应商")

        # Fetch effective prices and available stock for candidates on travel_date
        candidate_prices = []  # supplier_id, price, available, supplier_resource_id
        
        for cand in candidates:
            # Check ResourceInventory
            inv = await db.scalar(
                select(ResourceInventory).where(
                    ResourceInventory.supplier_resource_id == cand.id,
                    ResourceInventory.inventory_date == travel_date
                )
            )
            
            # Determine price
            price = cand.settlement_price
            if inv and inv.settlement_price is not None:
                price = inv.settlement_price
            
            available = 0
            if inv:
                available = max(0, inv.total_qty - inv.frozen_qty - inv.sold_qty)

            candidate_prices.append({
                "supplier_id": cand.supplier_id,
                "price": price if price is not None else Decimal(0),
                "available": available,
                "supplier_resource_id": cand.id,
            })

        # Filter: If manual selection exists, force it (no split)
        if line.resource_id in selections:
            target_sid = selections[line.resource_id]
            # Find in candidates
            found = [c for c in candidate_prices if c["supplier_id"] == target_sid]
            if not found:
                raise HTTPException(status_code=400, detail=f"所选供应商不适用于资源 {line.resource_id}")
            selected_cand = found[0]
            if selected_cand["available"] < qty_needed:
                raise HTTPException(status_code=400, detail=f"资源库存不足（资源 {line.resource_id}，供应商 {target_sid}）")
            allocations = [(selected_cand, qty_needed)]
        else:
            available_cands = [c for c in candidate_prices if c["available"] > 0]
            if not available_cands:
                raise HTTPException(status_code=400, detail=f"资源库存不足（资源 {line.resource_id}）")

            total_available = sum(c["available"] for c in available_cands)
            if total_available < qty_needed:
                raise HTTPException(status_code=400, detail=f"资源库存不足，无法拆分满足订单（资源 {line.resource_id}）")

            # Sort by price asc, then available asc (库存少优先), then supplier_id
            available_cands.sort(key=lambda x: (x["price"], x["available"], x["supplier_id"]))
            allocations = []
            remaining = qty_needed
            for cand in available_cands:
                if remaining <= 0:
                    break
                take = min(cand["available"], remaining)
                if take <= 0:
                    continue
                allocations.append((cand, take))
                remaining -= take
            if remaining > 0:
                raise HTTPException(status_code=400, detail=f"资源库存不足，无法拆分满足订单（资源 {line.resource_id}）")

        # Add to order resources (support split)
        for cand, take_qty in allocations:
            unit_cost = cand["price"]
            cost_amount = unit_cost * Decimal(take_qty)
            total_cost += cost_amount
            status_payload = {field: getattr(resource_input, field) for field in RESOURCE_STATUS_FIELDS}
            order_resources_data.append({
                "resource_id": line.resource_id,
                "supplier_id": cand["supplier_id"],
                "supplier_resource_id": cand["supplier_resource_id"],
                "quantity": take_qty,  # Resource qty for whole order (may be split)
                "settlement_price": unit_cost,
                "cost_amount": cost_amount,
                "travel_date": travel_date,
                "status_payload": status_payload,
                "composition_snapshot": composition_snapshot,
                "composition_snapshot_at": payload.paid_at if composition_snapshot else None,
            })

    if not primary_travel_date:
        raise HTTPException(status_code=400, detail="出行日期不能为空")
    active_price = await _active_price(db, payload.sku_id, payload.channel_id, primary_travel_date)
    if payload.sale_price is None:
        if not active_price or active_price.sale_price is None:
            raise HTTPException(status_code=400, detail="未找到该渠道有效价格，请填写销售金额或先配置价格")
    sale_price = Decimal(str(payload.sale_price if payload.sale_price is not None else active_price.sale_price))
    
    # Use calculated dynamic cost if payload doesn't provide it
    cost_price = Decimal(str(payload.cost_price)) if payload.cost_price is not None else (total_cost / Decimal(payload.quantity) if payload.quantity > 0 else 0)
    
    sale_amount, cost_amount, profit_amount = _calc_amounts(sale_price, cost_price, payload.quantity)
    if sale_amount <= 0:
        raise HTTPException(status_code=400, detail="销售金额必须大于 0")

    if payload.resource_items:
        for item in payload.resource_items:
            _validate_status_fields_for_resource(item, sale_amount, payload.quantity)
    else:
        _validate_status_fields_for_create(payload, sale_amount)

    # 3. Create Order
    order_travel_date = primary_travel_date or payload.travel_date
    order = Order(
        order_no=payload.order_no,
        channel_id=payload.channel_id,
        sku_id=payload.sku_id,
        product_id=product_id,
        travel_date=order_travel_date,
        quantity=payload.quantity,
        sale_price=sale_price,
        sale_amount=sale_amount,
        cost_price=cost_price,
        cost_amount=cost_amount,
        profit_amount=profit_amount,
        created_by=user.username,
        created_at=now_china(),
        remark=payload.remark,
        **{field: getattr(payload, field) for field in STATUS_FIELDS},
    )
    db.add(order)
    await db.flush()

    # 4. Freeze/Consume SKU Inventory (per distinct travel_date)
    issued_dates = sorted({item.travel_date for item in resource_input_map.values() if item.travel_date and item.is_issued})
    verified_dates = sorted({item.travel_date for item in resource_input_map.values() if item.travel_date and item.is_verified})
    try:
        for travel_date in issued_dates:
            await _freeze_inventory(
                db,
                payload.sku_id,
                payload.channel_id,
                travel_date,
                payload.quantity,
                user.username,
                order.id,
            )
        for travel_date in verified_dates:
            await _consume_inventory(
                db,
                payload.sku_id,
                travel_date,
                payload.quantity,
                user.username,
                order.id,
                "verify",
            )
    except HTTPException as e:
        raise e

    # 5. Create Order Resources and Freeze/Consume Resource Inventory
    for item in order_resources_data:
        status_payload = item.get("status_payload") or {}
        freeze_required = bool(status_payload.get("is_issued"))
        consume_required = bool(status_payload.get("is_verified"))
        if consume_required and not freeze_required:
            raise HTTPException(status_code=400, detail="核销/消耗前必须先出票/预约")
        # Create DB record
        or_rec = OrderResource(
            order_id=order.id,
            resource_id=item["resource_id"],
            supplier_id=item["supplier_id"],
            travel_date=item.get("travel_date"),
            quantity=item["quantity"],
            settlement_price=item["settlement_price"],
            cost_amount=item["cost_amount"],
            composition_snapshot=item.get("composition_snapshot"),
            composition_snapshot_at=item.get("composition_snapshot_at"),
            **status_payload,
        )
        db.add(or_rec)
        if not freeze_required and not consume_required:
            continue

        # Look up inventory again to lock
        inv = await db.scalar(
            select(ResourceInventory).where(
                ResourceInventory.supplier_resource_id == item["supplier_resource_id"],
                ResourceInventory.inventory_date == item.get("travel_date")
            ).with_for_update()
        )

        if not inv:
            raise HTTPException(status_code=400, detail=f"资源库存未初始化（资源 {item['resource_id']}，供应商 {item['supplier_id']}）")

        avail = inv.total_qty - inv.frozen_qty - inv.sold_qty
        if freeze_required:
            if avail < item["quantity"]:
                raise HTTPException(status_code=400, detail=f"资源库存不足（资源 {item['resource_id']}，供应商 {item['supplier_id']}）")
            inv.frozen_qty += item["quantity"]

        if consume_required:
            if inv.frozen_qty < item["quantity"]:
                raise HTTPException(status_code=400, detail="资源冻结库存不足")
            inv.frozen_qty -= item["quantity"]
            inv.sold_qty += item["quantity"]

        inv.updated_at = now_china().isoformat()
        db.add(inv)

    hist = OrderStatusHistory(
        order_id=order.id,
        before_status=None,
        after_status="created",
        operator=user.username,
        operated_at=now_china(),
        reason="created",
    )
    db.add(hist)
    await db.commit()
    await db.refresh(order)
    return OrderRead.model_validate(order)


@router.get("/orders/import-template")
async def download_import_template(
    db: DbSession,
    _: User = Depends(get_current_user),
):
    stmt = (
        select(SkuChannel, Sku, Spu, Channel)
        .join(Sku, Sku.id == SkuChannel.sku_id)
        .join(Spu, Spu.id == Sku.spu_id)
        .join(Channel, Channel.id == SkuChannel.channel_id)
        .limit(1)
    )
    row = (await db.execute(stmt)).first()
    if not row:
        raise HTTPException(status_code=400, detail="未找到可用的渠道 + SPU + SKU 绑定关系")

    sku_channel, sku, spu, channel = row
    now = now_china()
    travel_date = now.date()
    paid_at = now

    headers = ["订单号", "渠道", "SPU", "SKU", "数量", "销售金额", "出行日期", "支付时间"]
    for key, qty_label, amount_label in STATUS_TEMPLATE_FIELDS:
        meta = STATUS_IMPORT_META.get(key, {})
        headers.extend([
            meta.get("label", key),
            qty_label,
            amount_label,
            meta.get("time_label", "时间"),
        ])

    sample = [
        "SAMPLE-001",
        channel.channel_name,
        spu.name,
        sku.sku_name,
        "1",
        "100",
        travel_date.strftime("%Y-%m-%d"),
        paid_at.strftime("%Y-%m-%d %H:%M"),
    ]
    for _key, _qty_label, _amount_label in STATUS_TEMPLATE_FIELDS:
        sample.extend(["否", "", "", ""])

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(headers)
    ws.append(sample)

    guide = wb.create_sheet(title="Instructions")
    guide.append(["字段名", "是否必填", "格式/类型", "示例", "说明/约束"])
    guide.append(["订单号", "是", "文本", "ORD-20240202-001", "同一渠道内唯一（订单号 + 渠道 唯一）；重复会跳过"])
    guide.append(["渠道", "是", "文本（名称或ID）", "美团 / 12", "必须是系统已存在的渠道"])
    guide.append(["SPU", "是", "文本（名称/编码/ID）", "上海迪士尼门票 / SPU-001 / 3", "必须是系统已存在的 SPU"])
    guide.append(["SKU", "是", "文本（名称或ID）", "成人票 / 58", "必须是系统已存在的 SKU；且 SKU 必须属于所选 SPU"])
    guide.append(["数量", "是", "整数", "2", "必须大于 0"])
    guide.append(["销售金额", "是", "数字（总金额）", "199.00", "总销售金额；系统会按 数量 自动换算单价"])
    guide.append(["出行日期", "是", "日期 YYYY-MM-DD", "2026-02-02", "仅日期，不包含时间"])
    guide.append(["支付时间", "是", "时间 YYYY-MM-DD HH:mm", "2026-02-02 10:30", "必须填写，用于记录支付时间"])
    guide.append(["备注", "否", "文本", "客户补差价", "可选"])
    guide.append(["支付数量", "否", "整数", "2", "仅在需要记录支付数量时填写"])
    guide.append(["支付金额", "否", "数字", "199.00", "仅在需要记录支付金额时填写"])
    guide.append(["是否出票/预约（资源可用）", "否", "是/否", "是", "若为“是”，可填写“出票/预约时间”"])
    guide.append(["出票/预约数量", "否", "整数", "2", "仅当已出票/预约时可填"])
    guide.append(["出票/预约金额", "否", "数字", "199.00", "仅当已出票/预约时可填"])
    guide.append(["出票/预约时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-02 11:00", "可选"])
    guide.append(["是否核销/消耗", "否", "是/否", "否", "若为“是”，可填写“核销/消耗时间”"])
    guide.append(["核销/消耗数量", "否", "整数", "2", "仅当已核销/消耗时可填"])
    guide.append(["核销/消耗金额", "否", "数字", "199.00", "仅当已核销/消耗时可填"])
    guide.append(["核销/消耗时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-03 09:00", "可选"])
    guide.append(["是否支付后未核销/消耗全部退款", "否", "是/否", "否", "若为“是”，可填写“未核销/消耗退款时间”"])
    guide.append(["未核销/消耗退款数量", "否", "整数", "2", "仅当该退款发生时可填"])
    guide.append(["未核销/消耗退款金额", "否", "数字", "199.00", "仅当该退款发生时可填"])
    guide.append(["未核销/消耗退款时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-04 10:00", "可选"])
    guide.append(["是否支付后未出票/预约全部退款", "否", "是/否", "否", "若为“是”，可填写“未出票/预约退款时间”"])
    guide.append(["未出票/预约退款数量", "否", "整数", "2", "仅当该退款发生时可填"])
    guide.append(["未出票/预约退款金额", "否", "数字", "199.00", "仅当该退款发生时可填"])
    guide.append(["未出票/预约退款时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-04 10:30", "可选"])
    guide.append(["是否支付后已核销/消耗全部/部分退款", "否", "是/否", "否", "若为“是”，可填写“已核销/消耗退款时间”"])
    guide.append(["已核销/消耗退款数量", "否", "整数", "1", "仅当该退款发生时可填"])
    guide.append(["已核销/消耗退款金额", "否", "数字", "99.50", "仅当该退款发生时可填"])
    guide.append(["已核销/消耗退款时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-05 09:30", "可选"])
    guide.append(["是否支付后已出票/预约全部/部分退款", "否", "是/否", "否", "若为“是”，可填写“已出票/预约退款时间”"])
    guide.append(["已出票/预约退款数量", "否", "整数", "1", "仅当该退款发生时可填"])
    guide.append(["已出票/预约退款金额", "否", "数字", "99.50", "仅当该退款发生时可填"])
    guide.append(["已出票/预约退款时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-05 10:00", "可选"])
    guide.append(["是否完成", "否", "是/否", "否", "若为“是”，可填写“完成时间”"])
    guide.append(["完成数量", "否", "整数", "2", "仅当已完成时可填"])
    guide.append(["完成金额", "否", "数字", "199.00", "仅当已完成时可填"])
    guide.append(["完成时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-06 18:00", "可选"])
    guide.append(["是否完成后订单产生纠纷", "否", "是/否", "否", "若为“是”，可填写“纠纷时间”"])
    guide.append(["纠纷数量", "否", "整数", "1", "仅当产生纠纷时可填"])
    guide.append(["纠纷金额", "否", "数字", "50.00", "仅当产生纠纷时可填"])
    guide.append(["纠纷时间", "否", "时间 YYYY-MM-DD HH:mm", "2026-02-07 09:00", "可选"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = "order-import-template.xlsx"
    headers_resp = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.post("/orders/import")
async def import_orders(
    db: DbSession,
    user: User = Depends(require_roles(["admin", "operator", "csr"])),
    file: UploadFile = File(...),
):
    rows = await _read_upload_rows(file)
    if not rows:
        return {"created": 0, "skipped": 0, "errors": [], "total": 0}

    normalized_rows: list[tuple[int, dict]] = []
    sku_ids: set[int] = set()
    sku_names: set[str] = set()

    for row_idx, raw in rows:
        data = _normalize_row(raw)
        if not data:
            continue
        normalized_rows.append((row_idx, data))

        sku_raw = data.get("sku_id") or data.get("sku")
        sku_id_val = _parse_int(sku_raw)
        if sku_id_val:
            sku_ids.add(sku_id_val)
        else:
            sku_name_val = data.get("sku_name") or data.get("sku")
            if sku_name_val:
                sku_names.add(_normalize_text(sku_name_val))

    if not normalized_rows:
        return {"created": 0, "skipped": 0, "errors": [], "total": 0}

    channels = list(await db.scalars(select(Channel)))
    channel_by_id = {c.id: c for c in channels}
    channel_by_name = {_normalize_text(c.channel_name): c for c in channels}

    spus = list(await db.scalars(select(Spu)))
    spu_by_id = {s.id: s for s in spus}
    spu_by_name: dict[str, list[Spu]] = {}
    spu_by_code: dict[str, list[Spu]] = {}
    for s in spus:
        spu_by_name.setdefault(_normalize_text(s.name), []).append(s)
        if s.spu_code:
            spu_by_code.setdefault(_normalize_text(s.spu_code), []).append(s)

    sku_by_id: dict[int, Sku] = {}
    sku_by_name: dict[str, list[Sku]] = {}
    if sku_ids:
        skus = await db.scalars(select(Sku).where(Sku.id.in_(sku_ids)))
        for sku in skus:
            sku_by_id[sku.id] = sku
            sku_by_name.setdefault(_normalize_text(sku.sku_name), []).append(sku)
    if sku_names:
        skus = await db.scalars(select(Sku).where(func.lower(Sku.sku_name).in_(sku_names)))
        for sku in skus:
            sku_by_id.setdefault(sku.id, sku)
            sku_by_name.setdefault(_normalize_text(sku.sku_name), []).append(sku)

    sku_channel_map: dict[int, set[int]] = {}
    if sku_by_id:
        rows = await db.execute(
            select(SkuChannel.sku_id, SkuChannel.channel_id).where(
                SkuChannel.sku_id.in_(sku_by_id.keys()),
                SkuChannel.status == "active",
            )
        )
        for sku_id_val, channel_id_val in rows.all():
            sku_channel_map.setdefault(sku_id_val, set()).add(channel_id_val)

    created = 0
    skipped = 0
    errors: list[dict] = []

    for row_idx, data in normalized_rows:
        order_no = str(data.get("order_no") or "").strip()
        try:
            if not order_no:
                raise ValueError("订单号不能为空")

            channel_id_val = _parse_int(data.get("channel_id") or data.get("channel"))
            channel = channel_by_id.get(channel_id_val) if channel_id_val else None
            if not channel:
                channel_name_val = data.get("channel_name") or data.get("channel")
                if channel_name_val:
                    channel = channel_by_name.get(_normalize_text(channel_name_val))
            if not channel:
                raise ValueError("渠道不存在或未提供")

            spu_id_val = _parse_int(data.get("spu_id") or data.get("spu"))
            spu = spu_by_id.get(spu_id_val) if spu_id_val else None
            if not spu:
                spu_code_val = data.get("spu_code")
                spu_name_val = data.get("spu_name") or data.get("spu")
                if spu_code_val:
                    candidates = spu_by_code.get(_normalize_text(spu_code_val), [])
                    if len(candidates) > 1:
                        raise ValueError("SPU 编码重复，请提供 SPU ID")
                    if len(candidates) == 1:
                        spu = candidates[0]
                if not spu and spu_name_val:
                    candidates = spu_by_name.get(_normalize_text(spu_name_val), [])
                    if len(candidates) > 1:
                        raise ValueError("SPU 名称重复，请提供 SPU ID")
                    if len(candidates) == 1:
                        spu = candidates[0]
            if not spu:
                raise ValueError("SPU 不存在或未提供")

            sku_id_val = _parse_int(data.get("sku_id") or data.get("sku"))
            sku = sku_by_id.get(sku_id_val) if sku_id_val else None
            if not sku:
                sku_name_val = data.get("sku_name") or data.get("sku")
                if sku_name_val:
                    candidates = sku_by_name.get(_normalize_text(sku_name_val), [])
                    if spu:
                        candidates = [item for item in candidates if item.spu_id == spu.id]
                    if len(candidates) > 1:
                        raise ValueError("SKU 名称重复，请提供 SKU ID")
                    if len(candidates) == 1:
                        sku = candidates[0]
            if not sku:
                raise ValueError("SKU 不存在或未提供")
            if spu and sku.spu_id != spu.id:
                raise ValueError("SKU 不属于所选 SPU")
            if channel and sku_channel_map.get(sku.id) is not None:
                if channel.id not in sku_channel_map.get(sku.id, set()):
                    raise ValueError("渠道未绑定该 SKU")

            qty = _parse_int(data.get("quantity"))
            if not qty or qty <= 0:
                raise ValueError("数量必须大于 0")

            sale_amount_val = _parse_decimal(data.get("sale_amount"))
            sale_price_val = _parse_decimal(data.get("sale_price"))
            if sale_amount_val is None and sale_price_val is None:
                raise ValueError("销售金额不能为空")
            if sale_amount_val is not None and sale_amount_val <= 0:
                raise ValueError("销售金额必须大于 0")
            if sale_amount_val is None and sale_price_val is not None and sale_price_val <= 0:
                raise ValueError("销售单价必须大于 0")
            unit_price = None
            if sale_amount_val is not None:
                unit_price = (sale_amount_val / Decimal(qty)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            else:
                unit_price = sale_price_val

            travel_date = _parse_date(data.get("travel_date"))
            if not travel_date:
                raise ValueError("出行日期不能为空或格式错误")

            paid_at_raw = data.get("paid_at")
            if not paid_at_raw:
                raise ValueError("支付时间不能为空")
            paid_at = _parse_datetime(paid_at_raw)
            if not paid_at:
                raise ValueError("支付时间格式错误")

            remark = data.get("remark")

            status_payload: dict[str, object] = {}
            paid_qty = _parse_int(data.get("paid_qty"))
            if paid_qty is not None:
                status_payload["paid_qty"] = paid_qty
            paid_amount = _parse_decimal(data.get("paid_amount"))
            if paid_amount is not None:
                status_payload["paid_amount"] = float(paid_amount)
            has_paid_info = paid_at is not None or paid_qty is not None or paid_amount is not None
            status_payload = {
                "is_paid": bool(has_paid_info),
                "paid_at": paid_at,
                **status_payload,
            }

            for key in STATUS_IMPORT_KEYS:
                bool_val = _parse_bool(data.get(f"is_{key}"))
                qty_val = _parse_int(data.get(f"{key}_qty"))
                amount_val = _parse_decimal(data.get(f"{key}_amount"))
                at_val = _parse_datetime(data.get(f"{key}_at"))

                if bool_val is None:
                    if qty_val is not None or amount_val is not None or at_val is not None:
                        bool_val = True
                    else:
                        bool_val = False

                status_payload[f"is_{key}"] = bool_val
                if qty_val is not None:
                    status_payload[f"{key}_qty"] = qty_val
                if amount_val is not None:
                    status_payload[f"{key}_amount"] = float(amount_val)
                if at_val is not None:
                    status_payload[f"{key}_at"] = at_val

            if status_payload.get("is_verified") and not status_payload.get("is_issued"):
                raise ValueError("核销/消耗前必须先出票/预约")

            order_payload = OrderCreate(
                order_no=order_no,
                channel_id=channel.id,
                sku_id=sku.id,
                travel_date=travel_date,
                quantity=qty,
                sale_price=float(unit_price) if unit_price is not None else None,
                remark=remark,
                **status_payload,
            )

            await create_order(order_payload, db, user)
            created += 1
        except HTTPException as exc:
            await db.rollback()
            if str(exc.detail) == "该渠道下订单号已存在":
                skipped += 1
                continue
            errors.append({"row": row_idx, "order_no": order_no or None, "error": str(exc.detail)})
        except ValidationError as exc:
            await db.rollback()
            errors.append({"row": row_idx, "order_no": order_no or None, "error": "数据校验失败，请检查字段格式"})
        except Exception as exc:
            await db.rollback()
            errors.append({"row": row_idx, "order_no": order_no or None, "error": str(exc)})

    return {"created": created, "skipped": skipped, "errors": errors, "total": len(normalized_rows)}


@router.post("/orders/{order_id}/decision", response_model=OrderRead)
async def decide_order(
    payload: OrderDecision,
    db: DbSession,
    order_id: int = Path(..., ge=1),
    user: User = Depends(require_roles(["admin", "operator", "csr"])),
):
    order = await db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")
    order_resources = list(await db.scalars(select(OrderResource).where(OrderResource.order_id == order.id)))
    travel_dates = sorted({r.travel_date for r in order_resources if r.travel_date})
    if not travel_dates and order.travel_date:
        travel_dates = [order.travel_date]
    action = payload.action
    if action == "refund":
        action = "refund_unverified"

    action_meta_map = {
        "issue": "issue",
        "verify": "verified",
        "unverify": "unverify",
        "unissue": "unissue",
        "refund_unverified": "refund_unverified",
        "refund_unreserved": "refund_unreserved",
        "refund_verified": "refund_verified",
        "refund_reserved": "refund_reserved",
    }
    meta_key = action_meta_map.get(action)
    if payload.at is None:
        if meta_key:
            time_label = STATUS_META.get(meta_key, {}).get("time_label", "时间")
            raise HTTPException(status_code=400, detail=f"{time_label}不能为空")
        raise HTTPException(status_code=400, detail="时间不能为空")
    event_time = payload.at

    amount_val = _to_decimal(payload.amount)
    if amount_val is not None and amount_val < 0:
        label = STATUS_META.get(meta_key, {}).get("amount_label", "金额") if meta_key else "金额"
        raise HTTPException(status_code=400, detail=f"{label}不能小于 0")

    if action == "issue":
        target_qty, delta_qty = _resolve_qty(order, payload.qty, order.issued_qty)
        if delta_qty > 0:
            for travel_date in travel_dates:
                await _freeze_inventory(db, order.sku_id, order.channel_id, travel_date, delta_qty, user.username, order.id)
            await _apply_resource_inventory(db, order, delta_qty, user.username, "freeze")
        order.is_issued = True
        order.issued_at = event_time
        order.issued_qty = target_qty
        if payload.amount is not None:
            order.issued_amount = payload.amount
        for res in order_resources:
            res.is_issued = True
            res.issued_at = event_time
            res.issued_qty = target_qty
            if payload.amount is not None:
                res.issued_amount = payload.amount
    elif action == "unissue":
        if order.is_verified or any(res.is_verified for res in order_resources):
            raise HTTPException(status_code=400, detail="核销/消耗未取消，不能取消出票/预约")
        issued_qty = order.issued_qty
        if issued_qty is None:
            issued_qty = order.quantity if (order.is_issued or any(res.is_issued for res in order_resources)) else 0
        if issued_qty <= 0:
            raise HTTPException(status_code=400, detail="订单未出票/预约")
        issued_dates = sorted({r.travel_date for r in order_resources if r.travel_date and r.is_issued})
        if not issued_dates:
            issued_dates = travel_dates
        for travel_date in issued_dates:
            await _consume_inventory(db, order.sku_id, travel_date, issued_qty, user.username, order.id, "release")
        await _apply_resource_inventory(
            db,
            order,
            issued_qty,
            user.username,
            "release",
            resource_predicate=lambda r: r.is_issued,
        )
        order.is_issued = False
        order.issued_at = None
        order.issued_qty = None
        order.issued_amount = None
        for res in order_resources:
            if res.is_issued:
                res.is_issued = False
                res.issued_at = None
                res.issued_qty = None
                res.issued_amount = None
                res.issued_remark = None
    elif action == "verify":
        target_qty, delta_qty = _resolve_qty(order, payload.qty, order.verified_qty)
        if amount_val is not None:
            base_amount = _to_decimal(order.paid_amount) or _to_decimal(order.sale_amount)
            max_label = "支付金额" if order.paid_amount is not None else "销售金额"
            if base_amount is not None and amount_val > base_amount:
                raise HTTPException(status_code=400, detail=f"核销/消耗金额不能超过{max_label}")
        if delta_qty > 0:
            for travel_date in travel_dates:
                await _consume_inventory(db, order.sku_id, travel_date, delta_qty, user.username, order.id, "verify")
            await _apply_resource_inventory(db, order, delta_qty, user.username, "consume")
        order.is_verified = True
        order.verified_at = event_time
        order.verified_qty = target_qty
        if payload.amount is not None:
            order.verified_amount = payload.amount
        for res in order_resources:
            res.is_verified = True
            res.verified_at = event_time
            res.verified_qty = target_qty
            if payload.amount is not None:
                res.verified_amount = payload.amount
    elif action == "unverify":
        verified_qty = order.verified_qty
        if verified_qty is None:
            verified_qty = order.quantity if (order.is_verified or any(res.is_verified for res in order_resources)) else 0
        if verified_qty <= 0:
            raise HTTPException(status_code=400, detail="订单未核销/消耗")
        verified_dates = sorted({r.travel_date for r in order_resources if r.travel_date and r.is_verified})
        if not verified_dates:
            verified_dates = travel_dates
        for travel_date in verified_dates:
            await _unconsume_inventory(db, order.sku_id, travel_date, verified_qty, user.username, order.id)
        await _apply_resource_inventory(
            db,
            order,
            verified_qty,
            user.username,
            "unconsume",
            resource_predicate=lambda r: r.is_verified,
        )
        order.is_verified = False
        order.verified_at = None
        order.verified_qty = None
        order.verified_amount = None
        for res in order_resources:
            if res.is_verified:
                res.is_verified = False
                res.verified_at = None
                res.verified_qty = None
                res.verified_amount = None
                res.verified_remark = None
    elif action in {"refund_unverified", "refund_unreserved"}:
        field_qty = order.refund_unverified_qty if action == "refund_unverified" else order.refund_unreserved_qty
        target_qty, delta_qty = _resolve_qty(order, payload.qty, field_qty)
        if action == "refund_unverified":
            verified_qty = order.verified_qty if order.verified_qty is not None else (order.quantity if order.is_verified else 0)
            max_allowed = max(0, order.quantity - verified_qty)
            _ensure_unprocessed_refund_limit(target_qty, max_allowed, "核销/消耗")
            if amount_val is not None:
                base_amount = _to_decimal(order.paid_amount) or _to_decimal(order.sale_amount)
                max_label = "支付金额" if order.paid_amount is not None else "销售金额"
                if base_amount is not None and amount_val > base_amount:
                    raise HTTPException(status_code=400, detail=f"未核销/消耗退款金额不能超过{max_label}")
        else:
            issued_qty = order.issued_qty if order.issued_qty is not None else (order.quantity if order.is_issued else 0)
            max_allowed = max(0, order.quantity - issued_qty)
            _ensure_unprocessed_refund_limit(target_qty, max_allowed, "出票/预约")
            if amount_val is not None:
                base_amount = _to_decimal(order.paid_amount) or _to_decimal(order.sale_amount)
                max_label = "支付金额" if order.paid_amount is not None else "销售金额"
                if base_amount is not None and amount_val > base_amount:
                    raise HTTPException(status_code=400, detail=f"未出票/预约退款金额不能超过{max_label}")
        if delta_qty > 0:
            for travel_date in travel_dates:
                await _consume_inventory(db, order.sku_id, travel_date, delta_qty, user.username, order.id, "refund")
            await _apply_resource_inventory(db, order, delta_qty, user.username, "release")
        if action == "refund_unverified":
            order.is_refund_unverified = True
            order.refund_unverified_at = event_time
            order.refund_unverified_qty = target_qty
            if payload.amount is not None:
                order.refund_unverified_amount = payload.amount
            for res in order_resources:
                res.is_refund_unverified = True
                res.refund_unverified_at = event_time
                res.refund_unverified_qty = target_qty
                if payload.amount is not None:
                    res.refund_unverified_amount = payload.amount
        else:
            order.is_refund_unreserved = True
            order.refund_unreserved_at = event_time
            order.refund_unreserved_qty = target_qty
            if payload.amount is not None:
                order.refund_unreserved_amount = payload.amount
            for res in order_resources:
                res.is_refund_unreserved = True
                res.refund_unreserved_at = event_time
                res.refund_unreserved_qty = target_qty
                if payload.amount is not None:
                    res.refund_unreserved_amount = payload.amount
    elif action == "refund_verified":
        target_qty, delta_qty = _resolve_qty(order, payload.qty, order.refund_verified_qty)
        max_allowed = order.verified_qty if order.verified_qty is not None else (order.quantity if order.is_verified else 0)
        _ensure_refund_limit(target_qty, max_allowed, "核销/消耗")
        if amount_val is not None:
            base_amount = _to_decimal(order.verified_amount) or _to_decimal(order.sale_amount)
            max_label = "核销/消耗金额" if order.verified_amount is not None else "销售金额"
            if base_amount is not None and amount_val > base_amount:
                raise HTTPException(status_code=400, detail=f"已核销/消耗退款金额不能超过{max_label}")
        if delta_qty > 0:
            for travel_date in travel_dates:
                await _return_inventory(db, order.sku_id, travel_date, delta_qty, user.username, order.id)
            await _apply_resource_inventory(db, order, delta_qty, user.username, "return")
        order.is_refund_verified = True
        order.refund_verified_at = event_time
        order.refund_verified_qty = target_qty
        if payload.amount is not None:
            order.refund_verified_amount = payload.amount
        for res in order_resources:
            res.is_refund_verified = True
            res.refund_verified_at = event_time
            res.refund_verified_qty = target_qty
            if payload.amount is not None:
                res.refund_verified_amount = payload.amount
    elif action == "refund_reserved":
        target_qty, delta_qty = _resolve_qty(order, payload.qty, order.refund_reserved_qty)
        max_allowed = order.issued_qty if order.issued_qty is not None else (order.quantity if order.is_issued else 0)
        _ensure_refund_limit(target_qty, max_allowed, "出票/预约")
        if amount_val is not None:
            base_amount = _to_decimal(order.issued_amount) or _to_decimal(order.sale_amount)
            max_label = "出票/预约金额" if order.issued_amount is not None else "销售金额"
            if base_amount is not None and amount_val > base_amount:
                raise HTTPException(status_code=400, detail=f"已出票/预约退款金额不能超过{max_label}")
        if delta_qty > 0:
            for travel_date in travel_dates:
                await _return_inventory(db, order.sku_id, travel_date, delta_qty, user.username, order.id)
            await _apply_resource_inventory(db, order, delta_qty, user.username, "return")
        order.is_refund_reserved = True
        order.refund_reserved_at = event_time
        order.refund_reserved_qty = target_qty
        if payload.amount is not None:
            order.refund_reserved_amount = payload.amount
        for res in order_resources:
            res.is_refund_reserved = True
            res.refund_reserved_at = event_time
            res.refund_reserved_qty = target_qty
            if payload.amount is not None:
                res.refund_reserved_amount = payload.amount
    else:
        raise HTTPException(status_code=400, detail="不支持的操作")

    hist = OrderStatusHistory(
        order_id=order.id,
        before_status=None,
        after_status=action,
        operator=user.username,
        operated_at=now_china(),
        reason=payload.comment,
    )
    audit = AuditLog(
        table_name="order",
        record_id=order.id,
        operation="STATUS_CHANGE",
        diff_data={"action": action, "qty": payload.qty, "amount": payload.amount, "at": str(event_time)},
        operator=user.username,
        operated_at=now_china(),
        source=action,
    )
    db.add_all([order, hist, audit])
    await db.commit()
    await db.refresh(order)
    return OrderRead.model_validate(order)
