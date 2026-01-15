import io
import re
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.schemas.resource_attrs import (
    TicketAttrs, HotelAttrs, DiningAttrs, TransportAttrs
)

# 字段映射配置 (Field -> Chinese Column Name)
# 这里手动维护一份映射，确保列名对用户友好
ATTRS_MAPPING = {
    "门票": {
        "ticket_type": "票种(示例：成人票/儿童票)",
        "address": "地址(示例：景区正门)",
        "entrance_times": "入园次数(示例：1/unlimited)",
        "age_limit": '年龄限制(JSON示例：{"min":0,"max":100})',
        "includes": "包含内容",
        "excludes": "不包含内容",
        "earliest_entry_time": "最早入园时间(HH:MM)",
        "latest_entry_time": "最晚入园时间(HH:MM)",
        "advance_booking_days": "需提前预定天数(整数)",
        "advance_booking_time": '需提前预定时间(JSON示例：{"hours":2,"minutes":30})',
        "phone": "联系电话",
        "description": "详细介绍",
        "pickup_location": "取票地址",
        "available_after_issue": "出票后可用时间(示例：立即可用)",
        "required_traveler_info": "所需出行人信息(逗号分隔，示例：姓名,手机号)",
        "voucher_type": "凭证类型(逗号分隔，示例：二维码,身份证)",
        "purchase_limit": "限购规则",
        "refund_policy": "退票规则",
        "play_duration": "游玩时间(小时)",
        "additional_notes": "补充说明"
    },
    "酒店": {
        "room_type": "房型(示例：标准/豪华)",
        "bed_type": "床型(示例：大床/双床)",
        "hotel_type": "酒店类型(示例：经济型)",
        "phone": "联系电话",
        "address": "详细地址",
        "max_occupancy": "最大入住人数(整数)",
        "breakfast_included": "是否含早餐(是/否)",
        "star_rating": "酒店星级(示例：五星)",
        "cancellation_policy": "取消/退款政策",
        "purchase_limit": "限购规则",
        "required_traveler_info": "所需出行人信息(逗号分隔)",
        "description": "详细介绍",
        "check_in_time": "最早入住时间(HH:MM)",
        "check_out_time": "最晚退房时间(HH:MM)",
        "advance_booking_days": "提前预定天数(整数)",
        "parking": "停车场(示例：免费/付费/无)",
        "has_pickup_service": "接机/接站服务(是/否)",
        "has_24h_reception": "24小时前台(是/否)",
        "has_luggage_storage": "行李寄存(是/否)",
        "has_restaurant": "有餐厅(是/否)",
        "extra_services": "其他额外服务",
        "area": "面积(平米)",
        "special_settlement_rules": "特殊结算规则",
        "additional_notes": "补充说明"
    },
    "餐饮": {
        "meal_types": "餐饮类型(逗号分隔，示例：午餐,晚餐)",
        "dining_category": "餐饮分类(示例：正餐)",
        "restaurant_name": "餐厅名称",
        "restaurant_address": "餐厅地址",
        "phone": "联系电话",
        "opening_time": "营业开始时间(HH:MM)",
        "closing_time": "营业结束时间(HH:MM)",
        "reservation_required": "是否需要预定(是/否)",
        "additional_notes": "补充说明"
    },
    "交通": {
        "transport_type": "交通类型(示例：商务车)",
        "departure": "起点",
        "destination": "终点",
        "max_seats": "最大座位数",
        "duration": "行程时长",
        "additional_notes": "补充说明"
    },
    "组合": {} # 动态处理
}

TYPE_CN_TO_EN = {
    "酒店": "hotel",
    "门票": "ticket",
    "餐饮": "dining",
    "交通": "transport"
}

# 基础列定义
BASE_COLUMNS = [
    "资源名称(必填，唯一)", 
    "关联POI(必填，如：丽江古城)", 
    "供应商1(必填，如：携程)", 
    "结算价1(必填，数字)", 
    "供应商2(选填)", 
    "结算价2(选填)", 
    "供应商3(选填)", 
    "结算价3(选填)"
]

def generate_excel_template(resource_type: str, sub_types: Optional[str] = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{resource_type}导入模板"
    
    attr_columns = []
    
    if resource_type == "组合":
        # 组合类型特殊处理：根据 sub_types 生成带前缀的列
        # sub_types e.g. "酒店,门票"
        types_list = sub_types.split(",") if sub_types else []
        # 不需要用户手动填写 "包含类型" 列，后面解析时根据数据自动推断
        
        for t in types_list:
            t = t.strip()
            if t in ATTRS_MAPPING:
                t_map = ATTRS_MAPPING[t]
                for k, v in t_map.items():
                    # 生成 【类型】字段名 格式的列头
                    attr_columns.append(f"【{t}】{v}")
    else:
        # 获取该类型的特定字段
        attrs_map = ATTRS_MAPPING.get(resource_type, {})
        attr_columns = list(attrs_map.values())
    
    # 组合所有列头
    headers = BASE_COLUMNS + attr_columns
    
    # 写入列头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # 样式设置
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 简单列宽调整
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 25
        if resource_type == "组合" and "【" in header:
             ws.column_dimensions[column_letter].width = 30 # 加宽带前缀的列

    # === 添加示例行 ===
    # 基础列示例
    example_base = [
        "示例资源(上传前请删除此行)", # 资源名称
        "示例POI(如:丽江古城)",      # POI
        "携程",                       # 供应商1
        100,                          # 结算价1
        "美团",                       # 供应商2
        120,                          # 结算价2
        "",                           # 供应商3
        ""                            # 结算价3
    ]
    
    # 属性列示例值生成
    example_attrs = []
    
    if resource_type == "组合":
        # 为组合生成示例数据
        for header in attr_columns:
            val = ""
            # header format: 【类型】字段名
            if "JSON" in header:
                if "年龄" in header: val = '{"min":0,"max":100}'
                elif "时间" in header: val = '{"hours":1,"minutes":30}'
            elif "HH:MM" in header:
                val = "09:00"
            elif "是/否" in header:
                val = "是"
            elif "整数" in header:
                val = 1
            elif "类型" in header and "包含" not in header:
                    if "票种" in header: val = "成人票"
                    elif "房型" in header: val = "标准"
                    elif "床型" in header: val = "大床"
                    elif "交通" in header: val = "商务车"
            elif "逗号分隔" in header:
                val = "示例A,示例B"
            else:
                val = "示例填写"
            example_attrs.append(val)
    else:
        # 常规类型示例
        for col_name in attr_columns:
            val = ""
            if "JSON" in col_name:
                if "年龄" in col_name: val = '{"min":0,"max":100}'
                elif "时间" in col_name: val = '{"hours":1,"minutes":30}'
            elif "HH:MM" in col_name:
                val = "09:00"
            elif "是/否" in col_name:
                val = "是"
            elif "整数" in col_name:
                val = 1
            elif "类型" in col_name and "包含" not in col_name: 
                 if "票种" in col_name: val = "成人票"
                 elif "房型" in col_name: val = "标准"
                 elif "床型" in col_name: val = "大床"
                 elif "交通" in col_name: val = "商务车"
            elif "包含类型" in col_name:
                val = "酒店,门票"
            elif "逗号分隔" in col_name:
                val = "示例A,示例B"
            else:
                val = "示例填写"
            example_attrs.append(val)
        
    example_row = example_base + example_attrs
    
    for col_idx, val in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(color="808080", italic=True) # 灰色斜体提示是示例

    # 创建一个内存缓冲区
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def process_cell_value(val: Any, field_name: str) -> Any:
    """Helper to process cell value based on field name convention"""
    if val is not None:
        # 布尔值处理
        if isinstance(val, str) and field_name in [
            "breakfast_included", "has_pickup_service", "has_24h_reception", 
            "has_luggage_storage", "has_restaurant", "reservation_required"
        ]:
            if val.strip() in ["是", "Yes", "true", "True", "TRUE"]:
                return True
            elif val.strip() in ["否", "No", "false", "False", "FALSE"]:
                return False
        
        # 列表处理 (逗号分隔)
        if field_name in ["required_traveler_info", "voucher_type", "meal_types", "included_types"] and isinstance(val, str):
            # 分割并去除空白，过滤空字符串
            return [item.strip() for item in val.replace("，", ",").split(",") if item.strip()]
        
        # 字符串去除首尾空格
        if isinstance(val, str):
            return val.strip()
            
    return val

def parse_excel_data(file_content: bytes, resource_type: str, sub_types: Optional[str] = None) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    
    # 构建列名到字段的映射（仅针对非组合类型有效，组合类型需动态匹配）
    col_name_to_field = {}
    if resource_type != "组合":
        attrs_map = ATTRS_MAPPING.get(resource_type, {})
        col_name_to_field = {v: k for k, v in attrs_map.items()}
    
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
        
    data_list = []
    
    # 从第二行开始读取
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]: # 资源名称为空跳过
            continue
            
        row_data = {
            "row_idx": row_idx,
            "resource_name": str(row[0]).strip(),
            "poi_name": str(row[1]).strip() if row[1] else None,
            "suppliers": [],
            "attrs": {}
        }
        
        # 检查必填
        if not row_data["resource_name"]:
            raise ValueError(f"第 {row_idx} 行错误: 资源名称不能为空")
        if not row_data["poi_name"]:
            raise ValueError(f"第 {row_idx} 行错误: 关联POI不能为空")
            
        # 供应商解析 (Columns 2-7) [Index 2,3, 4,5, 6,7]
        for i in [2, 4, 6]:
            if i < len(row) and row[i]:
                s_name = str(row[i]).strip()
                try:
                    price_idx = i + 1
                    s_price = float(row[price_idx]) if price_idx < len(row) and row[price_idx] is not None else 0.0
                except (ValueError, TypeError):
                    raise ValueError(f"第 {row_idx} 行错误: 供应商结算价格式错误")
                row_data["suppliers"].append({"name": s_name, "price": s_price})
            
        if not row_data["suppliers"]:
             raise ValueError(f"第 {row_idx} 行错误: 至少需要填写供应商1")

        # Attrs 解析 (从 index 8 开始)
        for i in range(8, len(row)):
            if i >= len(headers): break
            
            header_name = headers[i]
            if not header_name: continue
            
            val = row[i]
            processed_val = None
            
            if resource_type == "组合":
                # 组合类型：匹配 【类型】字段名
                match = re.match(r"【(.*?)】(.*)", header_name)
                if match:
                    t_cn, f_cn = match.groups()
                    # 找到对应类型的字段key
                    t_map = ATTRS_MAPPING.get(t_cn, {})
                    # 反转映射: 中文名 -> key
                    rev_map = {v: k for k, v in t_map.items()}
                    f_key = rev_map.get(f_cn)
                    
                    if f_key:
                        t_en = TYPE_CN_TO_EN.get(t_cn)
                        if t_en:
                            if t_en not in row_data["attrs"]:
                                row_data["attrs"][t_en] = {}
                            
                            processed_val = process_cell_value(val, f_key)
                            # 只有值不为空时才设置，或者允许空值（取决于具体字段），这里都存入
                            row_data["attrs"][t_en][f_key] = processed_val

            else:
                # 普通类型
                field_name = col_name_to_field.get(header_name)
                if field_name:
                    processed_val = process_cell_value(val, field_name)
                    row_data["attrs"][field_name] = processed_val

        # 组合资源 included_types 处理
        if resource_type == "组合":
            if sub_types:
                # 优先使用传入的指定类型
                 row_data["attrs"]["included_types"] = [t.strip() for t in sub_types.split(",") if t.strip()]
            else:
                # 降级：自动推断
                inferred_types = []
                for type_cn, type_en in TYPE_CN_TO_EN.items():
                    # 检查该类型是否有任何数据被解析出来
                    if type_en in row_data["attrs"] and row_data["attrs"][type_en]:
                        has_data = False
                        for k, v in row_data["attrs"][type_en].items():
                            if v is not None and v != "":
                                has_data = True
                                break
                        if has_data:
                            inferred_types.append(type_cn)
                
                if inferred_types:
                    row_data["attrs"]["included_types"] = inferred_types

        data_list.append(row_data)
        
    return data_list
